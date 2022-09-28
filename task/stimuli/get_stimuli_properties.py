#!/usr/bin/env python
# coding: utf-8

# import necessary packages
import moviepy.editor as mpy
from os import walk
from os.path import join
import pandas as pd
from openpyxl import load_workbook
import sys

def create_dict(column_names):
    # create dictionaries with the elements in column_names as keys and empty lists as values
    data_dict = {}
    for col_name in column_names:
        data_dict[col_name] = []

    return data_dict

def get_gestures_and_actors_ind(in_dir):
    ind_gesture_start = len(in_dir)+4 # index of the start of the mentioning of 'GestureCode' in the filename
    ind_gesture_end = ind_gesture_start+4 # index of the end of the mentioning of 'GestureCode' in the filename
    ind_actor_start = ind_gesture_end+1 # index of the start of the mentioning of 'Actor' in the filename
    ind_actor_end = -4 # index of the end of the mentioning of 'Actor' in the filename

    return ind_gesture_start, ind_gesture_end, ind_actor_start, ind_actor_end

def get_interaction_type(left, right):
    # from the naming of the individual stimuli files determine which interaction type the output file will have
    if len(left) != len(right):
        return 'HR'
    else:
        return 'HH'

def get_pairing_type(left, right):
    # from the naming of the individual stimuli files determine which pairing type the output file will have
    if 'gg00' in [left, right]:
        return '1G'
    elif left != right:
        return '2G_DIFF'
    else:
        return '2G_SAME'

def get_stimuli_properties_basics(in_dir):
    column_names_basics = ['FilePath','FileName','FPS','Duration','Size']
    basics_dict = {}
    for col_name in column_names_basics:
        basics_dict[col_name] = []

    for subdir, dirs, files in walk(in_dir):
        for f in files:
            basics_dict['FileName'].append(f)

            file_path = join(subdir, f)
            basics_dict['FilePath'].append(file_path)

            clip = mpy.VideoFileClip(file_path,has_mask=True)

            duration = clip.duration
            basics_dict['Duration'].append(duration)

            fps = clip.fps
            basics_dict['FPS'].append(fps)

            size = clip.size
            basics_dict['Size'].append(size)

    return basics_dict

def create_sheet_singles(in_dir):
    column_names_singles = ['GestureCode','Actor']
    singles_dict = get_stimuli_properties_basics(in_dir)
    for col_name in column_names_singles:
        singles_dict[col_name] = []

    ind_gesture_start, ind_gesture_end, ind_actor_start, ind_actor_end = get_gestures_and_actors_ind(in_dir)

    for subdir, dirs, files in walk(in_dir):
        for f in files:
            file_path = join(subdir, f)
            gesture = file_path[ind_gesture_start:ind_gesture_end]
            singles_dict['GestureCode'].append(gesture)

            actor = file_path[ind_actor_start:ind_actor_end]
            singles_dict['Actor'].append(actor)

    # create sorted dataframes out of the dictionaries
    df_singles = pd.DataFrame(singles_dict).sort_values(by=['GestureCode', 'Actor']).reset_index(drop=True)
    df_singles.to_excel(f'{in_dir}.xlsx')

def create_sheet_dyads(in_dir, group):
    column_names_dyads = ['GestureCodeLeft','GestureCodeRight','ActorLeft','ActorRight','PairingType','InteractionType', 'Group']
    dyads_dict = get_stimuli_properties_basics(in_dir)
    for col_name in column_names_dyads:
        dyads_dict[col_name] = []

    _,_,_,ind_actor_end = get_gestures_and_actors_ind(in_dir)

    for subdir, dirs, files in walk(in_dir):
        for f in files:
            string_parts = f.split('_')

            dyads_dict['Group'].append(group)

            actor_right = string_parts[5][:ind_actor_end]
            dyads_dict['ActorRight'].append(actor_right)

            actor_left = string_parts[3]
            dyads_dict['ActorLeft'].append(actor_left)

            gesture_right = string_parts[4][1:] # disregard the first element because it is a prefixed 'r'
            dyads_dict['GestureCodeRight'].append(gesture_right)

            gesture_left = string_parts[2][1:] # disregard the first element because it is a prefixed 'l'
            dyads_dict['GestureCodeLeft'].append(gesture_left)

            interaction_type = get_interaction_type(actor_left, actor_right)
            dyads_dict['InteractionType'].append(interaction_type)

            pairing_type = get_pairing_type(gesture_left, gesture_right)
            dyads_dict['PairingType'].append(pairing_type)

    # create sorted dataframes out of the dictionaries
    df_dyads =  pd.DataFrame(dyads_dict).sort_values(by=['GestureCodeLeft', 'ActorLeft', 'GestureCodeRight', 'ActorRight']).reset_index(drop=True)
    df_dyads.to_excel(f'{in_dir}_{group}.xlsx')

def create_sheet_catchTrials(in_dir, group, total_frms):
    column_names_catchTrials = ['GestureCodeLeft','GestureCodeRight','ActorLeft','ActorRight','PairingType','InteractionType', 'CatchSide', 'CatchFrame', 'CatchTime','Group']
    catchTrials_dict = get_stimuli_properties_basics(in_dir)
    for col_name in column_names_catchTrials :
        catchTrials_dict[col_name] = []

    ind_actor_end = -4 # index of the end of the mentioning of 'Actor' in the filename

    for subdir, dirs, files in walk(in_dir):
        for f in files:
            string_parts = f.split('_')

            catchTrials_dict['Group'].append(group)

            catch_side = string_parts[2][0]
            catchTrials_dict['CatchSide'].append(catch_side)

            catch_frame = string_parts[2][1:]
            catchTrials_dict['CatchFrame'].append(catch_frame)

            catch_time = int(catch_frame)/total_frms
            catchTrials_dict['CatchTime'].append(catch_time)

            actor_right = string_parts[6][:ind_actor_end]
            catchTrials_dict['ActorRight'].append(actor_right)

            actor_left = string_parts[4]
            catchTrials_dict['ActorLeft'].append(actor_left)

            gesture_right = string_parts[5][1:] # disregard the first element because it is a prefixed 'r'
            catchTrials_dict['GestureCodeRight'].append(gesture_right)

            gesture_left = string_parts[3][1:] # disregard the first element because it is a prefixed 'l'
            catchTrials_dict['GestureCodeLeft'].append(gesture_left)

            interaction_type = get_interaction_type(actor_left, actor_right)
            catchTrials_dict['InteractionType'].append(interaction_type)

            pairing_type = get_pairing_type(gesture_left, gesture_right)
            catchTrials_dict['PairingType'].append(pairing_type)

    df_catchTrials =  pd.DataFrame(catchTrials_dict).sort_values(by=['GestureCodeLeft', 'ActorLeft', 'GestureCodeRight', 'ActorRight']).reset_index(drop=True)
    df_catchTrials.to_excel(f'{in_dir}_{group}.xlsx')
